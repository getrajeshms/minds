import sys
import os
import io
from datetime import date, datetime

import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data.assessment_questions import ASSESSMENT_CATEGORIES, get_ratings_for_type
from utils.assessment_engine import (
    get_active_questions,
    calculate_symptom_summaries,
    get_main_diagnosis,
)

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MindView — Mental Health Assessment",
    page_icon="🧠",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown("""
    <style>
    /* Hide default Streamlit chrome */
    #MainMenu, footer { visibility: hidden; }
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 860px; }

    /* Progress bar colour */
    .stProgress > div > div > div > div { background-color: #1e3a8a; }

    /* Subtle card wrapper */
    .card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.75rem 2rem;
        margin-bottom: 1.25rem;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    }

    /* Category badge */
    .cat-badge {
        display: inline-block;
        background: #dbeafe;
        color: #1d4ed8;
        padding: 4px 14px;
        border-radius: 999px;
        font-size: 0.82rem;
        font-weight: 600;
        letter-spacing: 0.02em;
        margin-bottom: 0.75rem;
    }

    /* Clinician note */
    .clinician-note {
        background: #fef3c7;
        border-left: 4px solid #f59e0b;
        padding: 10px 14px;
        border-radius: 0 6px 6px 0;
        color: #78350f;
        font-style: italic;
        font-size: 0.88rem;
        margin: 0.75rem 0 1rem;
    }

    /* Suicidal ideation alert */
    .suicidal-alert {
        background: #fee2e2;
        border: 1px solid #f87171;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        margin-bottom: 1.5rem;
        color: #991b1b;
    }

    /* Symptom row */
    .symptom-row {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 10px 16px;
        border-radius: 8px;
        margin-bottom: 6px;
    }

    /* Severity pill */
    .pill {
        font-size: 0.82rem;
        font-weight: 600;
        padding: 3px 12px;
        border-radius: 999px;
        border: 1px solid currentColor;
        background: #ffffff;
    }

    /* Report header */
    .report-header {
        background: #1e3a8a;
        color: white;
        border-radius: 12px 12px 0 0;
        padding: 1.5rem 2rem;
        text-align: center;
        margin-bottom: 0;
    }

    /* Section heading inside report */
    .section-heading {
        font-size: 1rem;
        font-weight: 700;
        color: #1e3a8a;
        border-bottom: 2px solid #dbeafe;
        padding-bottom: 6px;
        margin: 1.5rem 0 0.75rem;
    }

    /* Info row */
    .info-label { color: #6b7280; font-size: 0.8rem; margin: 0; }
    .info-value { font-weight: 600; margin: 0; font-size: 0.95rem; }

    /* Feature card */
    .feature-card {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        height: 100%;
    }

    /* Step indicator */
    .step-indicator {
        font-size: 0.78rem;
        color: #6b7280;
        text-align: right;
        margin-bottom: 0.25rem;
    }

    /* Main diagnosis box */
    .diagnosis-box {
        background: #eff6ff;
        border: 1px solid #bfdbfe;
        border-radius: 10px;
        padding: 1rem 1.5rem;
    }

    /* Disclaimer */
    .disclaimer {
        font-size: 0.8rem;
        color: #9ca3af;
        text-align: center;
        font-style: italic;
        margin-top: 1rem;
    }
    </style>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# State management
# ─────────────────────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        'step': 'welcome',
        'patient_info': {
            'name': '', 'age': '', 'gender': '',
            'unique_id': '', 'interviewer_name': '',
            'interview_date': str(date.today()),
        },
        'background_info': {
            'present_problems': '', 'duration': '', 'past_problems': '',
            'family_background': '', 'personal_social_background': '',
            'trauma': {'physical': False, 'emotional': False, 'sexual': False, 'neglect': False},
            'epilepsy': False, 'intellectual_disability': 'no',
        },
        'responses': {},
        'active_index': 0,
        'clinical_judgement': '',
        'treatment_given': '',
        'assistance_required': '',
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def reset_state():
    widget_keys = [k for k in list(st.session_state.keys())
                   if k.startswith('q_radio_') or k.startswith('q_notes_')]
    persistent = ['step', 'patient_info', 'background_info', 'responses',
                  'active_index', 'clinical_judgement', 'treatment_given', 'assistance_required']
    for k in persistent + widget_keys:
        if k in st.session_state:
            del st.session_state[k]
    init_state()


# ─────────────────────────────────────────────────────────────────────────────
# Shared components
# ─────────────────────────────────────────────────────────────────────────────
def render_progress_bar(pct: float):
    step = st.session_state.get('step', 'welcome')
    step_labels = {
        'patient-info': 'Step 1 of 4 — Patient Information',
        'background':   'Step 2 of 4 — Background',
        'assessment':   'Step 3 of 4 — Assessment',
        'clinical':     'Step 4 of 4 — Clinical Judgement',
        'report':       'Complete — Assessment Report',
    }
    label = step_labels.get(step, '')
    if step == 'assessment':
        qs = get_active_questions(st.session_state.responses)
        idx = st.session_state.active_index
        label = f"Step 3 of 4 — Assessment  ({idx + 1}/{len(qs)})"

    st.progress(min(pct / 100.0, 1.0))
    st.markdown(
        f"<p style='font-size:0.78rem;color:#6b7280;margin-top:-6px;'>{label}</p>",
        unsafe_allow_html=True,
    )


def logo_and_title(subtitle: str = ""):
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assets', 'a1intercept-logo.jpeg')
    if os.path.exists(logo_path):
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            st.image(logo_path, width=180)
    st.markdown(
        "<h1 style='text-align:center;color:#1e3a8a;margin-bottom:0.25rem;'>"
        "Mental Health Assessment Tool</h1>",
        unsafe_allow_html=True,
    )
    if subtitle:
        st.markdown(
            f"<p style='text-align:center;color:#6b7280;margin-top:0;'>{subtitle}</p>",
            unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# Welcome screen
# ─────────────────────────────────────────────────────────────────────────────
def render_welcome():
    logo_and_title(
        "A comprehensive semi-structured interview tool to assist healthcare professionals "
        "in making mental health assessments in primary care settings."
    )
    st.markdown("")

    col1, col2 = st.columns(2)
    features = [
        ("📋", "Structured Assessment",
         "Comprehensive evaluation across 16 mental health domains"),
        ("🧠", "Evidence-Based",
         "Based on the validated GMHAT/PC clinical assessment"),
        ("📄", "Detailed Reports",
         "Generate professional reports with symptom ratings and recommendations"),
        ("🔒", "Confidential",
         "All assessments remain private and are not stored externally"),
    ]
    for i, (icon, title, desc) in enumerate(features):
        with (col1 if i % 2 == 0 else col2):
            st.markdown(
                f"<div class='feature-card'>"
                f"<span style='font-size:1.5rem;'>{icon}</span>"
                f"<p style='font-weight:700;margin:6px 0 2px;'>{title}</p>"
                f"<p style='color:#6b7280;font-size:0.85rem;margin:0;'>{desc}</p>"
                f"</div>",
                unsafe_allow_html=True,
            )
            st.markdown("")

    with st.container(border=True):
        st.markdown("**Assessment Instructions**")
        st.markdown("""
- Please read the manual before administering the assessment
- Take account of all information, including reports from carers
- Rate symptoms the person has experienced in the **last month**
- You may ask supplementary questions to establish severity
""")

    col_l, col_m, col_r = st.columns([1, 1.2, 1])
    with col_m:
        if st.button("▶  Start New Assessment", type="primary", use_container_width=True):
            st.session_state.step = 'patient-info'
            st.rerun()

    st.markdown(
        "<p class='disclaimer'>This tool is an aid to healthcare professionals. "
        "It is not a substitute for detailed clinical assessment.<br>"
        "Powered by <strong>A1Intercept Technologies</strong></p>",
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Patient Info
# ─────────────────────────────────────────────────────────────────────────────
def render_patient_info():
    render_progress_bar(5)
    st.markdown("### 👤 Patient Information")

    pi = st.session_state.patient_info

    with st.form("patient_info_form"):
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("Patient Name *", value=pi.get('name', ''))
            age  = st.text_input("Age *", value=pi.get('age', ''))
            gender = st.selectbox(
                "Gender *",
                options=['', 'Male', 'Female', 'Other'],
                index=['', 'Male', 'Female', 'Other'].index(pi.get('gender', '')),
            )
        with col2:
            unique_id = st.text_input("Unique ID (optional)", value=pi.get('unique_id', ''))
            interviewer = st.text_input("Interviewer Name *", value=pi.get('interviewer_name', ''))
            try:
                default_date = date.fromisoformat(pi.get('interview_date', str(date.today())))
            except ValueError:
                default_date = date.today()
            interview_date = st.date_input("Interview Date *", value=default_date)

        st.markdown("")
        submitted = st.form_submit_button("Continue →", type="primary", use_container_width=True)

    if submitted:
        errors = []
        if not name.strip():         errors.append("Patient name is required.")
        if not age.strip():          errors.append("Age is required.")
        if not gender:               errors.append("Gender is required.")
        if not interviewer.strip():  errors.append("Interviewer name is required.")
        if errors:
            for e in errors:
                st.error(e)
        else:
            st.session_state.patient_info = {
                'name': name.strip(),
                'age': age.strip(),
                'gender': gender,
                'unique_id': unique_id.strip(),
                'interviewer_name': interviewer.strip(),
                'interview_date': str(interview_date),
            }
            st.session_state.step = 'background'
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# Background Form
# ─────────────────────────────────────────────────────────────────────────────
def render_background():
    render_progress_bar(10)
    st.markdown("### 📋 Background Information")

    bi = st.session_state.background_info

    with st.form("background_form"):
        present_problems = st.text_area(
            "Present Problems *",
            value=bi.get('present_problems', ''),
            height=100,
            placeholder="Describe the patient's current presenting problems…",
        )
        col1, col2 = st.columns(2)
        with col1:
            duration = st.text_input("Duration of Problems", value=bi.get('duration', ''))
        with col2:
            past_problems = st.text_input("Past Problems", value=bi.get('past_problems', ''))

        family_background = st.text_area(
            "Family Background",
            value=bi.get('family_background', ''),
            height=80,
        )
        personal_social = st.text_area(
            "Personal & Social Background",
            value=bi.get('personal_social_background', ''),
            height=80,
        )

        st.markdown("**Trauma History** *(check all that apply)*")
        trauma = bi.get('trauma', {})
        tc1, tc2 = st.columns(2)
        with tc1:
            physical  = st.checkbox("Physical trauma",  value=trauma.get('physical', False))
            emotional = st.checkbox("Emotional trauma", value=trauma.get('emotional', False))
        with tc2:
            sexual  = st.checkbox("Sexual trauma", value=trauma.get('sexual', False))
            neglect = st.checkbox("Neglect",        value=trauma.get('neglect', False))

        ec1, ec2 = st.columns(2)
        with ec1:
            epilepsy = st.checkbox("Epilepsy", value=bi.get('epilepsy', False))
        with ec2:
            id_opts = ['no', 'mild', 'severe']
            intellectual_disability = st.selectbox(
                "Intellectual Disability",
                options=id_opts,
                index=id_opts.index(bi.get('intellectual_disability', 'no')),
                format_func=lambda x: x.capitalize(),
            )

        st.markdown("")
        bc, nc = st.columns(2)
        with bc:
            back_btn = st.form_submit_button("← Back")
        with nc:
            next_btn = st.form_submit_button("Continue →", type="primary")

    def _save_bg():
        st.session_state.background_info = {
            'present_problems': present_problems.strip(),
            'duration': duration.strip(),
            'past_problems': past_problems.strip(),
            'family_background': family_background.strip(),
            'personal_social_background': personal_social.strip(),
            'trauma': {
                'physical': physical, 'emotional': emotional,
                'sexual': sexual, 'neglect': neglect,
            },
            'epilepsy': epilepsy,
            'intellectual_disability': intellectual_disability,
        }

    if back_btn:
        _save_bg()
        st.session_state.step = 'patient-info'
        st.rerun()

    if next_btn:
        if not present_problems.strip():
            st.error("Please describe the present problems.")
        else:
            _save_bg()
            st.session_state.step = 'assessment'
            st.session_state.active_index = 0
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# Assessment Question Card
# ─────────────────────────────────────────────────────────────────────────────
def render_assessment():
    responses = st.session_state.responses
    active_qs = get_active_questions(responses)
    total = len(active_qs)
    idx   = st.session_state.active_index

    # Guard — should not normally happen
    if idx >= total:
        st.session_state.step = 'clinical'
        st.rerun()
        return

    item = active_qs[idx]
    cat  = item['category']
    q    = item['question']

    progress_pct = 10 + (idx / total * 75) if total else 10
    render_progress_bar(progress_pct)

    # Category badge + question number
    col_badge, col_num = st.columns([3, 1])
    with col_badge:
        st.markdown(f"<span class='cat-badge'>{cat.icon} {cat.name}</span>", unsafe_allow_html=True)
    with col_num:
        st.markdown(f"<p class='step-indicator'>Question {idx + 1} of {total}</p>", unsafe_allow_html=True)

    # Question text
    st.markdown(f"<h3 style='margin-top:0.25rem;color:#111827;'>{q.question}</h3>", unsafe_allow_html=True)

    # Probing sub-questions
    if q.sub_questions:
        with st.expander("💬 Probing questions to explore with patient"):
            for sq in q.sub_questions:
                st.markdown(f"• {sq}")

    # Clinician note
    if q.clinician_note:
        st.markdown(
            f"<div class='clinician-note'>⚕️ {q.clinician_note}</div>",
            unsafe_allow_html=True,
        )

    st.markdown("---")

    # Rating
    ratings = get_ratings_for_type(q.rating_type)
    saved   = responses.get(q.id, {})
    saved_rating = saved.get('rating', 0)

    # Pre-populate widget key from saved response the first time
    radio_key = f"q_radio_{q.id}"
    if radio_key not in st.session_state:
        st.session_state[radio_key] = saved_rating

    selected_rating = st.radio(
        "**Select rating:**",
        options=[r['value'] for r in ratings],
        format_func=lambda v: f"{ratings[v]['label']}  —  {ratings[v]['description']}",
        key=radio_key,
    )

    # Notes
    notes_key = f"q_notes_{q.id}"
    if notes_key not in st.session_state:
        st.session_state[notes_key] = saved.get('notes', '')

    notes = st.text_area(
        "Clinician notes (optional):",
        key=notes_key,
        height=72,
    )

    st.markdown("")
    is_last  = (idx == total - 1)
    is_first = (idx == 0)

    bc, _, nc = st.columns([2, 0.4, 2])
    with bc:
        back_label = "← Background" if is_first else "← Back"
        if st.button(back_label, use_container_width=True):
            responses[q.id] = {'rating': selected_rating, 'notes': notes}
            if is_first:
                st.session_state.step = 'background'
            else:
                st.session_state.active_index -= 1
            st.rerun()

    with nc:
        next_label = "Complete Assessment →" if is_last else "Next →"
        if st.button(next_label, type="primary", use_container_width=True):
            responses[q.id] = {'rating': selected_rating, 'notes': notes}
            updated_qs = get_active_questions(responses)
            if st.session_state.active_index < len(updated_qs) - 1:
                st.session_state.active_index += 1
            else:
                st.session_state.step = 'clinical'
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# Clinical Judgement
# ─────────────────────────────────────────────────────────────────────────────
def render_clinical():
    render_progress_bar(90)
    st.markdown("### 🩺 Clinical Judgement")
    st.caption("Provide your overall clinical assessment before generating the report.")

    with st.form("clinical_form"):
        clinical_judgement = st.text_area(
            "Clinical Judgement / Summary *",
            value=st.session_state.get('clinical_judgement', ''),
            height=160,
            placeholder="Describe your overall clinical assessment, observations, and impressions…",
        )
        col1, col2 = st.columns(2)
        with col1:
            treatment_given = st.text_area(
                "Treatment Given",
                value=st.session_state.get('treatment_given', ''),
                height=100,
                placeholder="Medications, therapy, referrals…",
            )
        with col2:
            assistance_required = st.text_area(
                "Assistance Required",
                value=st.session_state.get('assistance_required', ''),
                height=100,
                placeholder="Support services, follow-up actions…",
            )

        st.markdown("")
        bc, nc = st.columns(2)
        with bc:
            back_btn = st.form_submit_button("← Back to Assessment")
        with nc:
            submit_btn = st.form_submit_button("Generate Report →", type="primary")

    def _save_clinical():
        st.session_state.clinical_judgement  = clinical_judgement.strip()
        st.session_state.treatment_given     = treatment_given.strip()
        st.session_state.assistance_required = assistance_required.strip()

    if back_btn:
        _save_clinical()
        active_qs = get_active_questions(st.session_state.responses)
        st.session_state.active_index = max(0, len(active_qs) - 1)
        st.session_state.step = 'assessment'
        st.rerun()

    if submit_btn:
        if not clinical_judgement.strip():
            st.error("Please provide a clinical judgement before generating the report.")
        else:
            _save_clinical()
            st.session_state.step = 'report'
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# Report
# ─────────────────────────────────────────────────────────────────────────────
SEVERITY_STYLE = {
    'No':       {'color': '#15803d', 'bg': '#dcfce7', 'icon': '✅'},
    'Mild':     {'color': '#a16207', 'bg': '#fef9c3', 'icon': '⚠️'},
    'Moderate': {'color': '#c2410c', 'bg': '#ffedd5', 'icon': '🔶'},
    'Severe':   {'color': '#b91c1c', 'bg': '#fee2e2', 'icon': '❌'},
}


def _fmt_date(iso: str) -> str:
    try:
        return datetime.fromisoformat(iso).strftime('%d %B %Y')
    except Exception:
        return iso


def render_report():
    render_progress_bar(100)

    pi        = st.session_state.patient_info
    bi        = st.session_state.background_info
    responses = st.session_state.responses
    summaries = calculate_symptom_summaries(responses)
    diagnosis = get_main_diagnosis(summaries)

    suicidal_resp  = responses.get('suicidal_ideation', {})
    suicidal_flag  = suicidal_resp.get('rating', 0) > 0
    completed_time = datetime.now().strftime('%d %B %Y, %H:%M')

    # ── Action buttons ────────────────────────────────────────────────────────
    btn1, btn2, btn3 = st.columns([1.4, 1.4, 1])
    with btn1:
        excel_bytes = _generate_excel(pi, bi, responses, summaries, diagnosis, suicidal_flag)
        fname = (
            f"GMHAT_{pi.get('name','').replace(' ','_')}"
            f"_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        st.download_button(
            "📊  Download Excel",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with btn2:
        if st.button("🖨️  Print / Save PDF", use_container_width=True):
            st.info("Use your browser's **File → Print** (Ctrl+P / ⌘+P) to save as PDF.")
    with btn3:
        if st.button("🔄  New Assessment", use_container_width=True):
            reset_state()
            st.rerun()

    # ── Suicidal ideation alert ───────────────────────────────────────────────
    if suicidal_flag:
        st.markdown(
            f"<div class='suicidal-alert'>"
            f"<strong>⚠️ Suicidal Ideation Flagged</strong><br>"
            f"This patient indicated thoughts of self-harm or ending their life "
            f"(score: {suicidal_resp['rating']}/3). "
            f"<strong>Immediate risk assessment and clinical follow-up is required.</strong>"
            f"</div>",
            unsafe_allow_html=True,
        )

    # ── Report header ─────────────────────────────────────────────────────────
    st.markdown(
        "<div class='report-header'>"
        "<h2 style='margin:0;font-size:1.4rem;'>GMHAT/PC Mental Health Assessment Report</h2>"
        f"<p style='margin:6px 0 0;opacity:0.8;font-size:0.88rem;'>"
        f"Interviewed by {pi.get('interviewer_name','')}  •  {completed_time}</p>"
        "</div>",
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        # Patient info
        st.markdown("<p class='section-heading'>Patient Information</p>", unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        for col, label, val in [
            (c1, "Name",      pi.get('name', '')),
            (c2, "Age",       pi.get('age', '')),
            (c3, "Gender",    pi.get('gender', '')),
            (c4, "Unique ID", pi.get('unique_id', 'N/A') or 'N/A'),
        ]:
            with col:
                st.markdown(
                    f"<p class='info-label'>{label}</p><p class='info-value'>{val}</p>",
                    unsafe_allow_html=True,
                )

        # Background
        st.markdown("<p class='section-heading'>Background Information</p>", unsafe_allow_html=True)
        st.markdown(
            f"<p class='info-label'>Present Problems</p>"
            f"<p class='info-value'>{bi.get('present_problems', 'Not specified')}</p>",
            unsafe_allow_html=True,
        )
        b1, b2 = st.columns(2)
        with b1:
            st.markdown(
                f"<p class='info-label'>Duration</p>"
                f"<p class='info-value'>{bi.get('duration','Not specified') or 'Not specified'}</p>",
                unsafe_allow_html=True,
            )
        with b2:
            st.markdown(
                f"<p class='info-label'>Past Problems</p>"
                f"<p class='info-value'>{bi.get('past_problems','None') or 'None'}</p>",
                unsafe_allow_html=True,
            )
        b3, b4 = st.columns(2)
        with b3:
            st.markdown(
                f"<p class='info-label'>Family Background</p>"
                f"<p class='info-value'>{bi.get('family_background','Not specified') or 'Not specified'}</p>",
                unsafe_allow_html=True,
            )
        with b4:
            trauma = bi.get('trauma', {})
            trauma_list = [k.capitalize() for k, v in trauma.items() if v]
            trauma_str  = ', '.join(trauma_list) if trauma_list else 'None reported'
            st.markdown(
                f"<p class='info-label'>Reported Trauma</p>"
                f"<p class='info-value'>{trauma_str}</p>",
                unsafe_allow_html=True,
            )
        b5, b6 = st.columns(2)
        with b5:
            st.markdown(
                f"<p class='info-label'>Epilepsy</p>"
                f"<p class='info-value'>{'Yes' if bi.get('epilepsy') else 'No'}</p>",
                unsafe_allow_html=True,
            )
        with b6:
            st.markdown(
                f"<p class='info-label'>Intellectual Disability</p>"
                f"<p class='info-value'>{bi.get('intellectual_disability','No').capitalize()}</p>",
                unsafe_allow_html=True,
            )

        # Symptoms
        st.markdown("<p class='section-heading'>Symptoms Based on Assessment</p>", unsafe_allow_html=True)
        st.caption("Domains marked 'Not indicated' were screened negative and not assessed further (GMHAT/PC branching logic).")

        for s in summaries:
            if s['skipped']:
                row_bg, pill_color, icon, label = '#f9fafb', '#6b7280', '—', 'Not indicated'
                score_html = ''
            else:
                cfg = SEVERITY_STYLE.get(s['severity'], SEVERITY_STYLE['No'])
                row_bg    = cfg['bg']
                pill_color = cfg['color']
                icon      = cfg['icon']
                label     = s['severity']
                score_html = f"<span style='color:#6b7280;font-size:0.82rem;margin-right:10px;'>Score: {s['score']}/{s['max_score']}</span>"

            st.markdown(
                f"<div class='symptom-row' style='background:{row_bg};'>"
                f"  <span style='font-weight:500;'>{icon} {s['category']}</span>"
                f"  <div style='display:flex;align-items:center;'>"
                f"    {score_html}"
                f"    <span class='pill' style='color:{pill_color};border-color:{pill_color};'>{label}</span>"
                f"  </div>"
                f"</div>",
                unsafe_allow_html=True,
            )

        # Main diagnosis
        st.markdown("<p class='section-heading'>Suggested Main Problem</p>", unsafe_allow_html=True)
        st.markdown(
            f"<div class='diagnosis-box'>"
            f"<p style='font-size:1.1rem;font-weight:700;color:#1e3a8a;margin:0;'>{diagnosis}</p>"
            f"</div>",
            unsafe_allow_html=True,
        )

        # Clinical judgement
        st.markdown("<p class='section-heading'>Clinical Judgement</p>", unsafe_allow_html=True)
        st.markdown(
            f"<p style='font-weight:500;'>{st.session_state.get('clinical_judgement','') or 'Not provided'}</p>",
            unsafe_allow_html=True,
        )

        # Treatment & assistance
        t1, t2 = st.columns(2)
        with t1:
            st.markdown("<p class='section-heading'>Treatment Given</p>", unsafe_allow_html=True)
            st.markdown(
                f"<p style='color:#374151;'>{st.session_state.get('treatment_given','') or 'Not specified'}</p>",
                unsafe_allow_html=True,
            )
        with t2:
            st.markdown("<p class='section-heading'>Assistance Required</p>", unsafe_allow_html=True)
            st.markdown(
                f"<p style='color:#374151;'>{st.session_state.get('assistance_required','') or 'Not specified'}</p>",
                unsafe_allow_html=True,
            )

        st.markdown(
            "<p class='disclaimer'>"
            "This assessment is presented as an aid to healthcare professionals for a quick mental health "
            "assessment. It is not a substitute for a detailed clinical assessment or making a diagnosis."
            "<br>Powered by <strong>A1Intercept Technologies</strong></p>",
            unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# Excel generation
# ─────────────────────────────────────────────────────────────────────────────
def _generate_excel(patient_info, background_info, responses, summaries, diagnosis, suicidal_flag) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    BLUE   = "1E3A8A"
    WHITE  = "FFFFFF"
    GREY   = "F3F4F6"
    bold14 = Font(bold=True, size=14, color=WHITE)
    bold11 = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    grey_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    def set_header(ws, row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12, color=BLUE)

    def col_width(ws, widths):
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    # ── Sheet 1: Patient & Background ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Patient Info"

    ws1.merge_cells("A1:B1")
    ws1["A1"] = "GMHAT/PC Mental Health Assessment Report"
    ws1["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws1["A1"].fill = hdr_fill
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.append([])
    set_header(ws1, 3, "Patient Information")
    for label, value in [
        ("Name",           patient_info.get('name', '')),
        ("Age",            patient_info.get('age', '')),
        ("Gender",         patient_info.get('gender', '')),
        ("Unique ID",      patient_info.get('unique_id', 'N/A') or 'N/A'),
        ("Interviewer",    patient_info.get('interviewer_name', '')),
        ("Interview Date", patient_info.get('interview_date', '')),
    ]:
        ws1.append([label, value])

    ws1.append([])
    set_header(ws1, ws1.max_row + 1, "Background Information")
    trauma = background_info.get('trauma', {})
    trauma_list = [k.capitalize() for k, v in trauma.items() if v]
    for label, value in [
        ("Present Problems",        background_info.get('present_problems', '')),
        ("Duration",                background_info.get('duration', '')),
        ("Past Problems",           background_info.get('past_problems', '')),
        ("Family Background",       background_info.get('family_background', '')),
        ("Personal/Social Bg",      background_info.get('personal_social_background', '')),
        ("Reported Trauma",         ', '.join(trauma_list) if trauma_list else 'None reported'),
        ("Epilepsy",                "Yes" if background_info.get('epilepsy') else "No"),
        ("Intellectual Disability", background_info.get('intellectual_disability', 'No').capitalize()),
    ]:
        ws1.append([label, value])

    col_width(ws1, {'A': 28, 'B': 55})
    for row in ws1.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Sheet 2: Symptoms ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Symptoms")
    headers = ["Category", "Score", "Max Score", "Severity", "Note"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font  = Font(bold=True, color=WHITE)
        cell.fill  = hdr_fill

    SEV_COLORS = {'No': 'D1FAE5', 'Mild': 'FEF9C3', 'Moderate': 'FFEDD5', 'Severe': 'FEE2E2'}
    for s in summaries:
        note = "Screening negative — not assessed further" if s['skipped'] else ""
        row_vals = [
            s['category'],
            s['score'],
            s['max_score'],
            "Not indicated" if s['skipped'] else s['severity'],
            note,
        ]
        ws2.append(row_vals)
        if not s['skipped']:
            fill_color = SEV_COLORS.get(s['severity'], GREY)
            for cell in ws2[ws2.max_row]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    col_width(ws2, {'A': 28, 'B': 10, 'C': 12, 'D': 16, 'E': 45})

    # ── Sheet 3: Detailed Responses ───────────────────────────────────────────
    ws3 = wb.create_sheet("Detailed Responses")
    ws3.append(["Category", "Question", "Rating (0-3)", "Notes"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = hdr_fill

    for cat in ASSESSMENT_CATEGORIES:
        for q in cat.questions:
            resp = responses.get(q.id)
            if resp is not None:
                ws3.append([cat.name, q.question, resp.get('rating', 0), resp.get('notes', '')])

    col_width(ws3, {'A': 25, 'B': 65, 'C': 14, 'D': 40})
    for row in ws3.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True)

    # ── Sheet 4: Clinical Summary ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Clinical Summary")
    ws4.merge_cells("A1:B1")
    ws4["A1"] = "Clinical Summary"
    ws4["A1"].font = Font(bold=True, size=13, color=WHITE)
    ws4["A1"].fill = hdr_fill
    ws4["A1"].alignment = Alignment(horizontal="center")

    ws4.append([])
    for label, value in [
        ("Suggested Main Problem",  diagnosis),
        ("Suicidal Ideation Flag",  "YES — requires follow-up" if suicidal_flag else "No"),
        ("Clinical Judgement",      st.session_state.get('clinical_judgement', '')),
        ("Treatment Given",         st.session_state.get('treatment_given', '')),
        ("Assistance Required",     st.session_state.get('assistance_required', '')),
    ]:
        ws4.append([label, value])

    col_width(ws4, {'A': 30, 'B': 70})
    for row in ws4.iter_rows(min_row=3):
        row[0].font = Font(bold=True)
        if len(row) > 1:
            row[1].alignment = Alignment(wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Main routing
# ─────────────────────────────────────────────────────────────────────────────
def main():
    inject_css()
    init_state()

    step = st.session_state.get('step', 'welcome')

    if step == 'welcome':
        render_welcome()
    elif step == 'patient-info':
        render_patient_info()
    elif step == 'background':
        render_background()
    elif step == 'assessment':
        render_assessment()
    elif step == 'clinical':
        render_clinical()
    elif step == 'report':
        render_report()


if __name__ == '__main__':
    main()
